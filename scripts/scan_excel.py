"""Scan 7 curriculum standard excel files and print structure."""
import os
import sys
import io
from openpyxl import load_workbook

OUT = open(r"C:\Users\user\OneDrive - 금성초등학교\바탕 화면\다채움 품질 제고사업 프로토타입 - 실동작\.claude\worktrees\distracted-blackwell\scripts\scan_out.txt", "w", encoding="utf-8")
_orig_print = print
def print(*a, **k):  # noqa
    _orig_print(*a, **k, file=OUT)

BASE = r"C:\Users\user\OneDrive - 금성초등학교\바탕 화면\다채움 품질 제고사업 프로토타입 - 실동작\교육과정표준체계_최종산출물_202412"

FILES = [
    "(KICE) 교육과정 표준체계_국어과_2412_최종.xlsx",
    "(KICE) 교육과정 표준체계_사회과_2412_최종.xlsx",
    "(KICE) 교육과정 표준체계_영어과_2412_최종.xlsx",
    "(KICE)) 교육과정 표준체계_실과및기술가정_2412_최종.xlsx",
    "(KOFAC) 교육과정 표준체계_과학과_2412_최종.xlsx",
    "(KOFAC) 교육과정 표준체계_정보과_2412_최종.xlsx",
    "3. 충북형_수학과_학습맵_계통도_KOFAC기준매핑_v2.xlsx",
]


def truncate(v, n=50):
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\r", " ").strip()
    return s[:n] + ("..." if len(s) > n else "")


def scan(path):
    print("=" * 90)
    print("FILE:", os.path.basename(path))
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print("  ERROR loading:", e)
        return
    print("  Sheets:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  --- Sheet '{sn}' rows={ws.max_row} cols={ws.max_column}")
    # dump first 6 rows of each sheet
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  >>> PREVIEW sheet='{sn}' <<<")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 6:
                break
            cells = [truncate(c, 50) for c in row]
            print(f"    R{i+1}:", cells)
    wb.close()


for f in FILES:
    p = os.path.join(BASE, f)
    if not os.path.exists(p):
        print("MISSING:", p)
        continue
    scan(p)
